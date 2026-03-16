from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from equipo import Equipo
from jugador import Jugador
from liga import Liga
from temporada import Temporada

try:
    import xlrd  # type: ignore
except Exception:
    xlrd = None


class Factoria:
    COLUMNAS_ESPERADAS = [
        "TEMPORADA", "LIGA", "EQUIPO", "JUGADOR", "PJUGADOS", "PCOMPLETOS",
        "PTITULAR", "PSUPLENTE", "MINUTOS", "LESIONES", "TARJETAS",
        "EXPULSIONES", "GOLES", "PENALTIES FALLADOS",
    ]

    PATRON_TEMPORADA = re.compile(r"^(\d{4})-(\d{2})$")

    @staticmethod
    def resolver_ruta_excel(ruta: str) -> str:
        candidatos: list[str] = []
        if os.path.isabs(ruta):
            candidatos.append(ruta)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
            candidatos.extend([
                ruta,
                os.path.join(base, ruta),
                os.path.join(base, "..", "data", ruta),
                os.path.join(os.getcwd(), ruta),
                os.path.join(os.getcwd(), "data", ruta),
                os.path.join(os.getcwd(), "EJERCICIO_4", "data", ruta),
            ])

        for candidato in candidatos:
            ruta_normalizada = os.path.abspath(candidato)
            if os.path.exists(ruta_normalizada):
                return ruta_normalizada
        return os.path.abspath(ruta)

    @staticmethod
    def _normalizar_texto(valor) -> str:
        if valor is None:
            return ""
        return str(valor).strip()

    @staticmethod
    def _a_entero(valor) -> int:
        if valor is None:
            return 0
        if isinstance(valor, bool):
            return int(valor)
        if isinstance(valor, int):
            return valor
        if isinstance(valor, float):
            return int(valor)
        texto = str(valor).strip()
        if texto == "":
            return 0
        texto = texto.replace(",", ".")
        try:
            return int(float(texto))
        except Exception:
            return 0

    @staticmethod
    def _extraer_anyo_inicio(temporada: str) -> int:
        encontrado = re.search(r"(\d{4})", temporada)
        if not encontrado:
            raise ValueError(f"No se pudo extraer el año inicial de la temporada: {temporada}")
        return int(encontrado.group(1))

    @staticmethod
    def _validar_formato_temporada(temporada: str) -> None:
        coincidencia = Factoria.PATRON_TEMPORADA.match(temporada)
        if coincidencia is None:
            raise ValueError(f"Formato de temporada inválido: {temporada}. Debe ser XXXX-YY")
        anio_inicio = int(coincidencia.group(1))
        anio_fin = int(coincidencia.group(2))
        esperado = (anio_inicio + 1) % 100
        if anio_fin != esperado:
            raise ValueError(f"Temporada incoherente en años consecutivos: {temporada}")

    @staticmethod
    def _leer_xlsx(ruta: str) -> list[dict[str, object]]:
        libro = load_workbook(filename=ruta, data_only=True, read_only=True)
        hoja = libro.active
        filas_iter = hoja.iter_rows(values_only=True)
        try:
            cabecera = next(filas_iter)
        except StopIteration:
            return []

        columnas = [Factoria._normalizar_texto(valor) for valor in cabecera]
        filas: list[dict[str, object]] = []

        for fila in filas_iter:
            if fila is None:
                continue
            registro: dict[str, object] = {}
            vacia = True
            for indice, columna in enumerate(columnas):
                valor = fila[indice] if indice < len(fila) else None
                registro[columna] = valor
                if valor not in (None, ""):
                    vacia = False
            if not vacia:
                filas.append(registro)

        libro.close()
        return filas

    @staticmethod
    def _leer_xls_con_xlrd(ruta: str) -> list[dict[str, object]]:
        if xlrd is None:
            raise ImportError("xlrd no está disponible")
        libro = xlrd.open_workbook(ruta)
        hoja = libro.sheet_by_index(0)
        columnas = [Factoria._normalizar_texto(hoja.cell_value(0, c)) for c in range(hoja.ncols)]
        filas: list[dict[str, object]] = []

        for fila_idx in range(1, hoja.nrows):
            registro: dict[str, object] = {}
            vacia = True
            for col_idx, columna in enumerate(columnas):
                valor = hoja.cell_value(fila_idx, col_idx)
                if valor != "":
                    vacia = False
                registro[columna] = valor
            if not vacia:
                filas.append(registro)

        return filas

    @staticmethod
    def _convertir_xls_a_xlsx_temporal(ruta: str) -> str:
        ejecutable = shutil.which("soffice") or shutil.which("libreoffice")
        if ejecutable is None:
            raise ImportError(
                "No se puede abrir .xls porque falta xlrd y no hay conversor local disponible. "
                "Instala xlrd o usa un archivo .xlsx."
            )

        carpeta_temporal = tempfile.mkdtemp(prefix="boletin4_xls_")
        comando = [
            ejecutable,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            carpeta_temporal,
            ruta,
        ]
        resultado = subprocess.run(comando, capture_output=True, text=True)
        if resultado.returncode != 0:
            raise RuntimeError(
                "No se pudo convertir el archivo .xls a .xlsx. "
                f"Salida: {resultado.stdout} {resultado.stderr}"
            )

        nombre_salida = Path(ruta).with_suffix(".xlsx").name
        ruta_convertida = os.path.join(carpeta_temporal, nombre_salida)
        if not os.path.exists(ruta_convertida):
            raise FileNotFoundError("La conversión a .xlsx terminó sin generar el archivo esperado.")
        return ruta_convertida

    @staticmethod
    def _leer_excel(ruta: str) -> list[dict[str, object]]:
        extension = os.path.splitext(ruta)[1].lower()
        if extension == ".xlsx":
            return Factoria._leer_xlsx(ruta)
        if extension == ".xls":
            if xlrd is not None:
                return Factoria._leer_xls_con_xlrd(ruta)
            ruta_convertida = Factoria._convertir_xls_a_xlsx_temporal(ruta)
            return Factoria._leer_xlsx(ruta_convertida)
        raise ValueError(f"Formato no soportado: {ruta}")

    @staticmethod
    def _validar_columnas(filas_crudas: list[dict[str, object]]) -> None:
        if not filas_crudas:
            raise ValueError("El Excel no contiene datos.")
        columnas_presentes = set(filas_crudas[0].keys())
        for columna in Factoria.COLUMNAS_ESPERADAS:
            if columna not in columnas_presentes:
                raise ValueError(f"Falta la columna obligatoria: {columna}")

    @staticmethod
    def _limpiar_filas(filas_crudas: list[dict[str, object]]) -> list[dict[str, object]]:
        filas_limpias: list[dict[str, object]] = []

        for registro in filas_crudas:
            nombre = Factoria._normalizar_texto(registro.get("JUGADOR"))
            if "TONTO" in nombre.upper():
                continue

            temporada = Factoria._normalizar_texto(registro.get("TEMPORADA"))
            Factoria._validar_formato_temporada(temporada)
            fila = {
                "TEMPORADA": temporada,
                "LIGA": Factoria._normalizar_texto(registro.get("LIGA")),
                "EQUIPO": Factoria._normalizar_texto(registro.get("EQUIPO")),
                "JUGADOR": nombre,
                "PJUGADOS": Factoria._a_entero(registro.get("PJUGADOS")),
                "PCOMPLETOS": Factoria._a_entero(registro.get("PCOMPLETOS")),
                "PTITULAR": Factoria._a_entero(registro.get("PTITULAR")),
                "PSUPLENTE": Factoria._a_entero(registro.get("PSUPLENTE")),
                "MINUTOS": Factoria._a_entero(registro.get("MINUTOS")),
                "LESIONES": Factoria._a_entero(registro.get("LESIONES")),
                "TARJETAS": Factoria._a_entero(registro.get("TARJETAS")),
                "EXPULSIONES": Factoria._a_entero(registro.get("EXPULSIONES")),
                "GOLES": Factoria._a_entero(registro.get("GOLES")),
                "PENALTIES FALLADOS": Factoria._a_entero(registro.get("PENALTIES FALLADOS")),
            }
            fila["anyo_inicio"] = Factoria._extraer_anyo_inicio(temporada)
            filas_limpias.append(fila)

        filas_limpias.sort(key=lambda fila: (fila["anyo_inicio"], fila["TEMPORADA"], fila["EQUIPO"], fila["JUGADOR"]))
        return filas_limpias

    @staticmethod
    def _asignar_jugador_id(filas_limpias: list[dict[str, object]]) -> None:
        agrupadas: dict[str, list[dict[str, object]]] = {}
        for fila in filas_limpias:
            nombre = str(fila["JUGADOR"])
            if nombre not in agrupadas:
                agrupadas[nombre] = []
            agrupadas[nombre].append(fila)

        for nombre, grupo in agrupadas.items():
            grupo.sort(key=lambda fila: (fila["anyo_inicio"], fila["EQUIPO"], fila["TEMPORADA"]))
            cluster = 1
            fila_previa: dict[str, object] | None = None

            for fila in grupo:
                if fila_previa is not None:
                    gap = int(fila["anyo_inicio"]) - int(fila_previa["anyo_inicio"])
                    cambia_cluster_por_gap = gap > 5
                    cambia_cluster_por_misma_temporada = (
                        fila["TEMPORADA"] == fila_previa["TEMPORADA"] and fila["EQUIPO"] != fila_previa["EQUIPO"]
                    )
                    if cambia_cluster_por_gap or cambia_cluster_por_misma_temporada:
                        cluster += 1

                fila["jugador_id"] = f"{nombre}#{cluster}"
                fila_previa = fila

    @staticmethod
    def _validar_fila(fila: dict[str, object], partidos_temporada: int) -> None:
        campos_numericos = [
            "PJUGADOS", "PCOMPLETOS", "PTITULAR", "PSUPLENTE", "MINUTOS", "LESIONES",
            "TARJETAS", "EXPULSIONES", "GOLES", "PENALTIES FALLADOS",
        ]
        for campo in campos_numericos:
            if int(fila[campo]) < 0:
                raise ValueError(f"Cantidad negativa en {campo} para {fila['JUGADOR']} ({fila['TEMPORADA']}).")

        if int(fila["PCOMPLETOS"]) > int(fila["PTITULAR"]):
            raise ValueError(f"PCOMPLETOS > PTITULAR en {fila['JUGADOR']} ({fila['TEMPORADA']}).")

        if int(fila["PJUGADOS"]) != int(fila["PSUPLENTE"]) + int(fila["PTITULAR"]):
            raise ValueError(f"PJUGADOS != PSUPLENTE + PTITULAR en {fila['JUGADOR']} ({fila['TEMPORADA']}).")

        if int(fila["MINUTOS"]) > int(fila["PJUGADOS"]) * 90:
            raise ValueError(f"MINUTOS > PJUGADOS*90 en {fila['JUGADOR']} ({fila['TEMPORADA']}).")

        if int(fila["PJUGADOS"]) > partidos_temporada:
            raise ValueError(f"PJUGADOS > partidos de la temporada en {fila['JUGADOR']} ({fila['TEMPORADA']}).")

    @staticmethod
    def _validar_datos(filas_limpias: list[dict[str, object]]) -> None:
        equipos_por_temporada: dict[str, set[str]] = {}
        for fila in filas_limpias:
            temporada = str(fila["TEMPORADA"])
            equipos_por_temporada.setdefault(temporada, set()).add(str(fila["EQUIPO"]))

        partidos_por_temporada = {
            temporada: len(equipos) * (len(equipos) - 1) if len(equipos) > 1 else 0
            for temporada, equipos in equipos_por_temporada.items()
        }

        for fila in filas_limpias:
            temporada = str(fila["TEMPORADA"])
            Factoria._validar_fila(fila, partidos_por_temporada[temporada])

    @staticmethod
    def _crear_liga(filas_limpias: list[dict[str, object]]) -> Liga:
        liga = Liga()

        for fila in filas_limpias:
            jugador = Jugador(
                jugador_id=str(fila["jugador_id"]),
                nombre=str(fila["JUGADOR"]),
                equipo=str(fila["EQUIPO"]),
                temporada=str(fila["TEMPORADA"]),
                anyo_inicio=int(fila["anyo_inicio"]),
                pjugados=int(fila["PJUGADOS"]),
                pcompletos=int(fila["PCOMPLETOS"]),
                ptitular=int(fila["PTITULAR"]),
                psuplente=int(fila["PSUPLENTE"]),
                minutos=int(fila["MINUTOS"]),
                lesiones=int(fila["LESIONES"]),
                tarjetas=int(fila["TARJETAS"]),
                expulsiones=int(fila["EXPULSIONES"]),
                goles=int(fila["GOLES"]),
                pen_fallados=int(fila["PENALTIES FALLADOS"]),
            )

            id_temporada = jugador.temporada
            if id_temporada not in liga.temporadas:
                liga.agregar_temporada(Temporada(identificador=id_temporada))
            temporada = liga.temporadas[id_temporada]

            if jugador.equipo not in temporada.equipos:
                temporada.agregar_equipo(Equipo(nombre=jugador.equipo, temporada=id_temporada))
            equipo = temporada.equipos[jugador.equipo]
            equipo.agregar_jugador(jugador)

        return liga

    @staticmethod
    def cargar_excel(ruta: str) -> Liga:
        ruta_resuelta = Factoria.resolver_ruta_excel(ruta)
        if not os.path.exists(ruta_resuelta):
            raise FileNotFoundError(f"No se encontró el Excel: {ruta_resuelta}")

        filas_crudas = Factoria._leer_excel(ruta_resuelta)
        Factoria._validar_columnas(filas_crudas)
        filas_limpias = Factoria._limpiar_filas(filas_crudas)
        Factoria._asignar_jugador_id(filas_limpias)
        Factoria._validar_datos(filas_limpias)
        return Factoria._crear_liga(filas_limpias)


class FactoriaFutbol(Factoria):
    """Alias compatible con versiones previas del proyecto."""
